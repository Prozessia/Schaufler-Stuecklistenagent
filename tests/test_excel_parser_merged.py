"""Ticket DATA-002: Excel parser — merged cells, sheet selection, header detection,
and position anchor.

Tests build minimal workbooks in tmp_path using openpyxl so they are fully
deterministic (no real files needed).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.ingestion.excel_parser import parse_excel


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------


def _wb_path(tmp_path: Path, name: str = "test.xlsx") -> Path:
    return tmp_path / name


# ---------------------------------------------------------------------------
# Test 1: vertically merged position cell — all three rows carry the value
# ---------------------------------------------------------------------------


def test_merged_cells_vertical_propagation(tmp_path: Path) -> None:
    """A position cell merged over 3 rows must appear in all 3 data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "BOM"

    # Row 1 — header
    ws["A1"] = "Position"
    ws["B1"] = "Benennung"
    ws["C1"] = "Werkstoff"

    # Row 2 — position "10" spans rows 2-4 (merged vertically)
    ws["A2"] = "10"
    ws["B2"] = "Formplatte"
    ws["C2"] = "1.2343"

    ws["B3"] = "Schieber"
    ws["C3"] = "1.2344"

    ws["B4"] = "Kern"
    ws["C4"] = "AlSi9Cu3"

    # Merge A2:A4 (vertical)
    ws.merge_cells("A2:A4")

    path = _wb_path(tmp_path)
    wb.save(path)

    result = parse_excel(path)

    # We should get 3 data rows and all should carry "10" in the Position column
    assert len(result.rows) == 3, f"Expected 3 rows, got {len(result.rows)}: {result.rows}"
    for idx, row in enumerate(result.rows):
        assert row.get("Position") == "10", (
            f"Row {idx} Position is {row.get('Position')!r}, expected '10'"
        )


# ---------------------------------------------------------------------------
# Test 2: cover sheet first → BOM sheet second — BOM sheet must be chosen
# ---------------------------------------------------------------------------


def test_sheet_selection_prefers_bom_over_cover(tmp_path: Path) -> None:
    """Sheet with BOM header keywords beats a plain text cover sheet."""
    wb = openpyxl.Workbook()

    # Sheet 1 — cover page with lots of text but no BOM keywords
    cover = wb.active
    assert cover is not None
    cover.title = "Deckblatt"
    for r in range(1, 12):
        cover[f"A{r}"] = f"Metadaten Zeile {r}"
        cover[f"B{r}"] = f"Wert {r}"
        cover[f"C{r}"] = f"Info {r}"

    # Sheet 2 — actual BOM
    bom_ws = wb.create_sheet("Daten")
    bom_ws["A1"] = "Pos"
    bom_ws["B1"] = "Benennung"
    bom_ws["C1"] = "Menge"
    bom_ws["D1"] = "Werkstoff"
    for r in range(2, 8):
        bom_ws[f"A{r}"] = r - 1
        bom_ws[f"B{r}"] = f"Teil {r - 1}"
        bom_ws[f"C{r}"] = 1
        bom_ws[f"D{r}"] = "1.2343"

    path = _wb_path(tmp_path)
    wb.save(path)

    result = parse_excel(path)

    assert result.metadata.get("sheet_name") == "Daten", (
        f"Expected sheet 'Daten', got {result.metadata.get('sheet_name')!r}"
    )
    assert len(result.rows) >= 5, f"Expected >=5 data rows, got {len(result.rows)}"


# ---------------------------------------------------------------------------
# Test 3: header in row 3 under two meta rows — keyword score wins
# ---------------------------------------------------------------------------


def test_header_detection_keyword_beats_text_count(tmp_path: Path) -> None:
    """Header row with BOM keywords wins even if earlier rows have more text cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "BOM"

    # Row 1 — many text cells but no BOM keywords (meta block)
    ws["A1"] = "Firma"
    ws["B1"] = "Adresse"
    ws["C1"] = "Telefon"
    ws["D1"] = "Fax"
    ws["E1"] = "Datum"
    ws["F1"] = "Version"
    ws["G1"] = "Erstellt von"

    # Row 2 — another meta row
    ws["A2"] = "SCHAUFLER Tooling"
    ws["B2"] = "Musterstraße 1"
    ws["C2"] = "+49 123 456"
    ws["D2"] = "+49 123 457"
    ws["E2"] = "2024-01-01"
    ws["F2"] = "1.0"
    ws["G2"] = "Max Mustermann"

    # Row 3 — actual header with BOM keywords (fewer cells but keyword hits)
    ws["A3"] = "Position"
    ws["B3"] = "Benennung"
    ws["C3"] = "Menge"
    ws["D3"] = "Werkstoff"

    # Data rows
    for r in range(4, 7):
        ws[f"A{r}"] = r - 3
        ws[f"B{r}"] = f"Bauteil {r - 3}"
        ws[f"C{r}"] = 2
        ws[f"D{r}"] = "1.2343"

    path = _wb_path(tmp_path)
    wb.save(path)

    result = parse_excel(path)

    assert result.metadata.get("header_row") == 3, (
        f"Expected header_row=3, got {result.metadata.get('header_row')}"
    )
    assert "Position" in result.headers, f"Headers: {result.headers}"
    assert "Benennung" in result.headers, f"Headers: {result.headers}"


# ---------------------------------------------------------------------------
# Test 4: expected_position_count and raw_pdf_positions correctly set
# ---------------------------------------------------------------------------


def test_position_anchor_set_from_data_rows(tmp_path: Path) -> None:
    """expected_position_count and raw_pdf_positions are derived from data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "BOM"

    ws["A1"] = "Pos"
    ws["B1"] = "Benennung"
    ws["C1"] = "Werkstoff"

    positions = ["1", "2", "3", "4", "5"]
    for r, pos in enumerate(positions, start=2):
        ws[f"A{r}"] = int(pos)
        ws[f"B{r}"] = f"Bauteil {pos}"
        ws[f"C{r}"] = "1.2343"

    path = _wb_path(tmp_path)
    wb.save(path)

    result = parse_excel(path)

    assert result.expected_position_count == 5, (
        f"Expected 5, got {result.expected_position_count}"
    )
    assert result.raw_pdf_positions == ["1", "2", "3", "4", "5"], (
        f"raw_pdf_positions: {result.raw_pdf_positions}"
    )
