"""Surfacing of lossless non-data (footer/header/note) flags.

Covers both reviewer-facing surfaces:
* the API result (RowResult.non_data / non_data_reasons), and
* the Excel export (a non-destructive cell comment on the flagged row).

Both must stay lossless: every row is still present.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from src.api.job_store import Job
from src.api.result_builder import build_job_result
from src.export.excel_exporter import export_to_excel
from src.mapping.schema_registry import load_schema
from src.scoring.audit_trail import BomAuditTrail, CellAudit
from src.scoring.threshold_manager import TrafficLight

_DATA_START_ROW = 7


def _audit_with_flag() -> BomAuditTrail:
    schema = load_schema()
    pos = schema.fields[0]  # Detail Number, column A
    cells = [
        CellAudit(
            row_index=i,
            target_field=pos.name,
            target_column=pos.column,
            transformed_value=str(10 * (i + 1)),
            classification=TrafficLight.GREEN,
            final_score=0.95,
        )
        for i in range(3)
    ]
    return BomAuditTrail(
        source_file="demo.pdf",
        customer="Demo",
        cells=cells,
        expected_position_count=3,
        green_count=3,
        total_scored=3,
        # row 1 looks like a footer/note
        non_data_row_flags={1: ["NO_POSITION", "FOOTER_OR_HEADER_TEXT"]},
    )


def test_api_result_exposes_non_data_flag():
    audit = _audit_with_flag()
    job = Job(
        job_id="job-x",
        filename="demo.pdf",
        filepath=Path("demo.pdf"),
        customer="Demo",
        status="completed",
        audit=audit,
    )
    result = build_job_result(
        job.job_id, job, template_path=Path("missing-template.xlsx")
    )

    by_index = {r.row_index: r for r in result.rows}
    # lossless: all three rows present
    assert set(by_index) == {0, 1, 2}
    # only the flagged row carries the advisory flag
    assert by_index[1].non_data is True
    assert "FOOTER_OR_HEADER_TEXT" in by_index[1].non_data_reasons
    assert by_index[0].non_data is False
    assert by_index[2].non_data is False


def test_excel_export_marks_flagged_row_with_comment(tmp_path: Path):
    audit = _audit_with_flag()
    out = tmp_path / "out.xlsx"

    export_to_excel(audit, out, add_audit_sheet=False)

    wb = openpyxl.load_workbook(out)
    ws = wb["Stückliste"] if "Stückliste" in wb.sheetnames else wb.active

    # flagged row 1 is the 2nd written data row
    flagged_cell = ws.cell(row=_DATA_START_ROW + 1, column=1)
    assert flagged_cell.comment is not None
    assert "FOOTER_OR_HEADER_TEXT" in flagged_cell.comment.text

    # the other data rows are present and NOT commented (lossless, no false marks)
    assert ws.cell(row=_DATA_START_ROW + 0, column=1).comment is None
    assert ws.cell(row=_DATA_START_ROW + 2, column=1).comment is None
    assert ws.cell(row=_DATA_START_ROW + 0, column=1).value is not None
    assert ws.cell(row=_DATA_START_ROW + 2, column=1).value is not None
