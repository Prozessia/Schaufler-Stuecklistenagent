"""R3 — reviewer row exclusion: audited, reflected in export, zero-loss preserved.

Excluding a row is an explicit user action, NOT silent data loss. These tests pin:
  * exclusion mutates the audit (logged + counts recomputed),
  * the exporter skips excluded rows AND drops their identities from the guard,
  * a genuinely dropped (non-excluded) row still raises ZeroDataLossError,
  * restoring an excluded row reverses everything,
  * build_job_result hides excluded rows from the result payload.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

from src.api.cell_edits import apply_row_exclusions
from src.core.exceptions import ZeroDataLossError
from src.export import excel_exporter
from src.export.excel_exporter import export_to_excel
from src.mapping.schema_registry import load_schema
from src.scoring.audit_trail import BomAuditTrail, CellAudit
from src.scoring.threshold_manager import TrafficLight

_TEMPLATE = excel_exporter._CONFIG_DIR / "target_template.xlsx"


def _band_audit(n_rows: int) -> BomAuditTrail:
    """Audit with one position cell per row, each on a distinct RB-1 band id."""
    schema = load_schema()
    field = schema.fields[0]  # position field, column A
    cells = [
        CellAudit(
            row_index=i,
            source_row_id=f"p0:b{i:04d}",
            target_field=field.name,
            target_column=field.column,
            transformed_value=str(i + 1),
            classification=TrafficLight.GREEN,
        )
        for i in range(n_rows)
    ]
    keys = [c.source_row_id for c in cells]
    return BomAuditTrail(
        source_file="test.pdf",
        customer="K",
        cells=cells,
        expected_row_keys=keys,
        expected_position_count=n_rows,
        guard_basis="row_band_set",
    )


def _save_spy(monkeypatch: pytest.MonkeyPatch) -> MagicMock:
    real_wb = openpyxl.load_workbook(_TEMPLATE)
    spy = MagicMock()
    real_wb.save = spy  # type: ignore[method-assign]
    monkeypatch.setattr(
        excel_exporter.openpyxl, "load_workbook", lambda *a, **k: real_wb
    )
    return spy


# ---------------------------------------------------------------------------
# apply_row_exclusions
# ---------------------------------------------------------------------------


def test_exclude_marks_row_logged_and_recounts() -> None:
    audit = _band_audit(3)
    audit.green_count = 3
    audit.total_scored = 3

    changed = apply_row_exclusions(audit, [1], excluded=True)

    assert changed == [1]
    assert audit.excluded_rows == [1]
    assert audit.green_count == 2  # excluded row no longer counted
    assert audit.total_scored == 2
    assert len(audit.exclusion_log) == 1
    entry = audit.exclusion_log[0]
    assert entry.row_index == 1
    assert entry.source_row_id == "p0:b0001"
    assert entry.excluded_at  # ISO timestamp stamped


def test_exclude_is_idempotent_and_unknown_rows_ignored() -> None:
    audit = _band_audit(3)
    apply_row_exclusions(audit, [1], excluded=True)
    changed = apply_row_exclusions(audit, [1, 999], excluded=True)

    assert changed == []  # already excluded; 999 unknown
    assert audit.excluded_rows == [1]
    assert len(audit.exclusion_log) == 1


def test_restore_reverses_exclusion() -> None:
    audit = _band_audit(3)
    apply_row_exclusions(audit, [1], excluded=True)

    changed = apply_row_exclusions(audit, [1], excluded=False)

    assert changed == [1]
    assert audit.excluded_rows == []
    assert audit.exclusion_log == []
    assert audit.green_count == 3


# ---------------------------------------------------------------------------
# Exporter guard
# ---------------------------------------------------------------------------


def test_export_skips_excluded_row_without_data_loss_error(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    spy = _save_spy(monkeypatch)
    audit = _band_audit(3)
    apply_row_exclusions(audit, [1], excluded=True)

    # Band p0:b0001 is legitimately gone -> guard must NOT raise.
    export_to_excel(audit, tmp_path / "out.xlsx", add_audit_sheet=False)

    spy.assert_called_once()


def test_export_still_raises_for_genuine_loss_alongside_exclusion(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    spy = _save_spy(monkeypatch)
    audit = _band_audit(3)
    apply_row_exclusions(audit, [1], excluded=True)
    # Now SILENTLY drop a different, non-excluded row's cells (real data loss).
    audit.cells = [c for c in audit.cells if c.row_index != 2]

    with pytest.raises(ZeroDataLossError, match="Zeilen-Bändern fehlen"):
        export_to_excel(audit, tmp_path / "out.xlsx", add_audit_sheet=False)

    spy.assert_not_called()


def test_export_writes_only_active_rows(tmp_path: Path) -> None:
    audit = _band_audit(3)
    apply_row_exclusions(audit, [1], excluded=True)
    out = tmp_path / "out.xlsx"

    export_to_excel(audit, out, add_audit_sheet=False, colour_cells=False)

    wb = openpyxl.load_workbook(out)
    ws = wb[excel_exporter._SHEET_NAME]
    start = excel_exporter._DATA_START_ROW
    col = 1  # position column A
    written = [
        ws.cell(row=start + i, column=col).value for i in range(3)
    ]
    # Two active rows written ("1","3"); third data row empty.
    assert str(written[0]) == "1"
    assert str(written[1]) == "3"
    assert written[2] is None


# ---------------------------------------------------------------------------
# Result payload
# ---------------------------------------------------------------------------


def test_build_job_result_hides_excluded_rows() -> None:
    from src.api.job_store import Job
    from src.api.result_builder import build_job_result

    audit = _band_audit(3)
    apply_row_exclusions(audit, [1], excluded=True)
    job = Job(
        job_id="job-1",
        filename="test.pdf",
        filepath=Path("test.pdf"),
        customer="K",
        status="completed",
        audit=audit,
    )

    result = build_job_result("job-1", job)

    row_indices = {row.row_index for row in result.rows}
    assert row_indices == {0, 2}
    assert result.excluded_rows == [1]
    assert result.total_rows == 2
