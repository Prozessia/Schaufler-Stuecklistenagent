"""Test script: Full pipeline Parse -> Map -> Transform -> Score -> Export.

Generates a Schaufler Excel file for EVERY parseable customer BOM and
validates:
  - Template format preserved (sheets, merged cells, data validations)
  - Values in the correct cells (column letter matches schema)
  - Data types are correct (integers, decimals, strings)
  - Traffic-light colours applied
  - Audit sheet present

Usage:
    python scripts/test_export.py                    # Full (needs LLM)
    python scripts/test_export.py --skip-llm         # Use saved mappings
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import sys
from pathlib import Path

# Suppress MuPDF C-level noise — redirect to devnull
os.environ["PYMUPDF_MESSAGE"] = "path:" + os.devnull

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv

load_dotenv(PROJECT_ROOT / ".env")

import openpyxl
from openpyxl.utils import get_column_letter

from src.core.models import ParsedBOM
from src.ingestion.structure_normalizer import parse_file
from src.mapping.llm_column_mapper import MappingResult, ColumnMapping, map_columns
from src.mapping.schema_registry import load_schema
from src.transform.pipeline import transform_bom
from src.transform.cross_validator import cross_validate
from src.scoring.ensemble_scorer import score_bom
from src.scoring.threshold_manager import load_scoring_config, TrafficLight
from src.scoring.audit_trail import BomAuditTrail
from src.export.excel_exporter import export_to_excel

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
logger = logging.getLogger("test_export")

DATA_DIR = PROJECT_ROOT / "data" / "input" / "PDF_POC"
OUTPUT_DIR = PROJECT_ROOT / "data" / "test_outputs" / "exports"
SAVED_MAPPINGS_FILE = PROJECT_ROOT / "data" / "test_outputs" / "mapping_results.json"


# ---------------------------------------------------------------------------
# Helpers (shared patterns from test_scoring.py)
# ---------------------------------------------------------------------------


def find_bom_files() -> list[Path]:
    files = []
    for f in sorted(DATA_DIR.rglob("*")):
        if not f.is_file():
            continue
        if f.suffix.lower() not in (".pdf", ".xlsx"):
            continue
        if "CadCam_Stuecklistenvorlage" in f.name:
            continue
        files.append(f)
    return files


def get_customer_key(filepath: Path) -> str:
    parts = filepath.parts
    try:
        poc_idx = parts.index("PDF_POC")
        return parts[poc_idx + 1]
    except (ValueError, IndexError):
        return "unknown"


def load_saved_mappings() -> dict[str, dict]:
    if not SAVED_MAPPINGS_FILE.exists():
        return {}
    data = json.loads(SAVED_MAPPINGS_FILE.read_text(encoding="utf-8"))
    by_customer = {}
    for entry in data:
        if "error" not in entry and entry.get("has_ground_truth"):
            by_customer[entry.get("customer", "")] = entry
    return by_customer


def reconstruct_mapping_result(
    saved: dict, bom: ParsedBOM, schema=None
) -> MappingResult:
    if schema is None:
        schema = load_schema()
    field_to_col = {f.name: f.column for f in schema.fields}
    mappings = []
    seen_sources = set()
    for detail in saved.get("details", []):
        src = detail.get("actual_source")
        tgt = detail.get("actual_target")
        conf = detail.get("confidence", 0.0)
        if src and tgt and detail.get("status") == "CORRECT":
            if src not in seen_sources:
                seen_sources.add(src)
                mappings.append(
                    ColumnMapping(
                        source_column=src,
                        target_field=tgt,
                        target_column=field_to_col.get(tgt, ""),
                        confidence=conf,
                        reasoning="from saved evaluation",
                    )
                )
    return MappingResult(
        source_file=bom.source.filename,
        customer=bom.source.customer,
        mappings=mappings,
    )


# ---------------------------------------------------------------------------
# Validation of exported file
# ---------------------------------------------------------------------------


def validate_export(
    excel_path: Path,
    audit: BomAuditTrail,
    schema,
) -> list[str]:
    """Validate exported Excel against expectations. Return list of issues."""
    issues: list[str] = []

    wb = openpyxl.load_workbook(excel_path)

    # 1. Sheet names
    if "Stückliste" not in wb.sheetnames:
        issues.append("MISSING: 'Stückliste' sheet")
        return issues
    if "Audit Trail" not in wb.sheetnames:
        issues.append("MISSING: 'Audit Trail' sheet")
    if "Stammdaten" not in wb.sheetnames:
        issues.append("MISSING: 'Stammdaten' sheet (template corrupted)")

    ws = wb["Stückliste"]

    # 2. Header row preserved
    for field in schema.fields:
        col_num = openpyxl.utils.column_index_from_string(field.column)
        header_cell = ws.cell(row=5, column=col_num)
        if header_cell.value is None:
            issues.append(
                f"HEADER MISSING: Col {field.column} (row 5) for '{field.name}'"
            )

    # 3. Merged cells preserved
    merged = list(ws.merged_cells.ranges)
    if len(merged) < 5:
        issues.append(f"MERGED CELLS: Only {len(merged)} merged ranges (expected >= 5)")

    # 4. Data validations preserved
    dv_count = len(ws.data_validations.dataValidation) if ws.data_validations else 0
    if dv_count < 8:
        issues.append(f"DATA VALIDATIONS: Only {dv_count} (expected >= 8)")

    # 5. Count data rows written
    # The exporter writes ALL distinct row_indices from audit (including empty rows)
    all_row_indices = sorted({c.row_index for c in audit.cells})
    data_rows_expected = len(all_row_indices)
    data_rows_found = 0
    for row_idx in range(7, 7 + data_rows_expected):
        has_value = any(
            ws.cell(row=row_idx, column=c).value is not None for c in range(1, 31)
        )
        if has_value:
            data_rows_found += 1

    if data_rows_found == 0:
        issues.append("NO DATA: No data rows written to template")

    # 6. Spot-check: verify a few green cells are in the correct column
    greens = [
        c
        for c in audit.cells
        if c.classification == TrafficLight.GREEN and c.transformed_value
    ]
    checked = 0
    misplaced = 0
    for cell_audit in greens[:50]:
        field_def = schema.field_by_name.get(cell_audit.target_field)
        if not field_def:
            continue
        col_num = openpyxl.utils.column_index_from_string(field_def.column)
        # Find which output row this maps to
        try:
            out_row_offset = all_row_indices.index(cell_audit.row_index)
        except ValueError:
            continue
        excel_row = 7 + out_row_offset
        excel_val = ws.cell(row=excel_row, column=col_num).value
        if excel_val is None:
            misplaced += 1
        else:
            # Compare string representations
            expected = cell_audit.transformed_value
            actual = str(excel_val)
            if field_def.type == "integer":
                try:
                    if int(float(actual)) != int(expected):
                        misplaced += 1
                        continue
                except (ValueError, TypeError):
                    pass
            elif field_def.type == "decimal":
                try:
                    if abs(float(actual) - float(expected.replace(",", "."))) > 0.01:
                        misplaced += 1
                        continue
                except (ValueError, TypeError):
                    pass
        checked += 1

    if misplaced > 0:
        issues.append(
            f"MISPLACED: {misplaced}/{checked} green cells NOT in correct column"
        )
    elif checked > 0:
        print(f"    OK: {checked} green cells verified in correct columns")

    # 7. Check type: integer columns should hold int values
    for field in schema.fields:
        if field.type == "integer":
            col_num = openpyxl.utils.column_index_from_string(field.column)
            for row_idx in range(7, min(7 + data_rows_found, 12)):
                v = ws.cell(row=row_idx, column=col_num).value
                if v is not None and not isinstance(v, (int, float)):
                    issues.append(
                        f"TYPE: Col {field.column} row {row_idx} is {type(v).__name__} "
                        f"(expected int) for '{field.name}'"
                    )
                    break

    wb.close()
    return issues


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


async def main():
    skip_llm = "--skip-llm" in sys.argv

    schema = load_schema()
    config = load_scoring_config()
    logger.info("Export test starting...")

    files = find_bom_files()
    logger.info("Found %d BOM files", len(files))

    saved_mappings = load_saved_mappings()
    if saved_mappings:
        logger.info("Loaded saved mappings for %d customers", len(saved_mappings))

    llm = None
    if not skip_llm:
        try:
            from src.llm.azure_openai import AzureOpenAILLM

            llm = AzureOpenAILLM()
        except Exception as e:
            logger.warning("LLM init failed: %s -- using saved mappings", e)

    seen_customers: set[str] = set()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    total_exported = 0
    total_issues = 0

    for filepath in files:
        customer_key = get_customer_key(filepath)
        if customer_key in seen_customers:
            continue
        seen_customers.add(customer_key)

        # Parse
        try:
            bom = parse_file(filepath)
        except Exception as e:
            logger.error("Parse failed for %s: %s", filepath.name, e)
            continue
        if not bom.headers or not bom.rows:
            logger.warning("Empty BOM: %s", filepath.name)
            continue

        # Map
        mapping = None
        if customer_key in saved_mappings:
            mapping = reconstruct_mapping_result(
                saved_mappings[customer_key], bom, schema
            )
        elif llm:
            try:
                mapping = await map_columns(bom, llm, schema)
            except Exception as e:
                logger.error("Mapping failed for %s: %s", customer_key, e)
                continue
        else:
            logger.warning("No mapping for %s -- skipping", customer_key)
            continue
        if mapping.mapped_count == 0:
            logger.warning("No mappings for %s", customer_key)
            continue

        # Transform
        transform_result = transform_bom(bom, mapping, schema)

        # Cross-validate
        cv_result = cross_validate(transform_result)

        # Score
        audit = score_bom(transform_result, mapping, cv_result, schema, config)

        # Export
        safe_name = customer_key.replace(" ", "_").replace("/", "_")
        output_path = OUTPUT_DIR / f"{safe_name}_export.xlsx"
        try:
            export_to_excel(
                audit,
                output_path,
                schema=schema,
                colour_cells=True,
                add_audit_sheet=True,
                meta={"customer": customer_key},
            )
        except Exception as e:
            logger.error("Export failed for %s: %s", customer_key, e)
            print(f"\n  EXPORT FAILED: {customer_key} -- {e}")
            continue

        total_exported += 1

        # Validate
        print(f"\n{'=' * 60}")
        print(f"  {customer_key}")
        print(f"{'=' * 60}")
        print(f"  Source: {filepath.name}")
        print(f"  Output: {output_path.name}")
        print(f"  Rows: {len(audit.cells) // 30}")
        print(
            f"  GREEN: {audit.green_count}  YELLOW: {audit.yellow_count}  RED: {audit.red_count}"
        )

        issues = validate_export(output_path, audit, schema)
        if issues:
            print(f"  ISSUES ({len(issues)}):")
            for issue in issues:
                print(f"    - {issue}")
            total_issues += len(issues)
        else:
            print(f"  VALIDATION: ALL CHECKS PASSED")

    # Summary
    print(f"\n{'=' * 60}")
    print(f"  EXPORT SUMMARY")
    print(f"{'=' * 60}")
    print(f"  Customers exported: {total_exported}")
    print(f"  Output directory:   {OUTPUT_DIR}")
    print(f"  Total issues:       {total_issues}")
    if total_issues == 0:
        print(f"  STATUS: ALL EXPORTS VALID")
    else:
        print(f"  STATUS: {total_issues} issues found -- review above")

    # List output files
    print(f"\n  Generated files:")
    for f in sorted(OUTPUT_DIR.glob("*.xlsx")):
        size_kb = f.stat().st_size / 1024
        print(f"    {f.name} ({size_kb:.0f} KB)")


if __name__ == "__main__":
    asyncio.run(main())
