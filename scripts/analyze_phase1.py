"""Phase 1: Comprehensive data analysis of all BOM files."""

import json
import sys
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
from openpyxl.cell.cell import MergedCell

BASE = Path(r"c:\Users\douioui\Documents\privat\Schaufler-Stücklistenagent")
INPUT_DIR = BASE / "data" / "input" / "PDF_POC"


def analyze_excel(filepath: Path) -> dict:
    result = {
        "type": "excel",
        "file": filepath.name,
        "size_kb": round(filepath.stat().st_size / 1024, 1),
    }
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        result["sheets"] = wb.sheetnames
        sheets_data = {}
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows_data = []
            merged = [str(m) for m in ws.merged_cells.ranges] if ws.merged_cells else []
            max_rows_to_read = min(ws.max_row or 1, 50)
            for row in ws.iter_rows(
                min_row=1, max_row=max_rows_to_read, values_only=False
            ):
                row_vals = []
                for cell in row:
                    if isinstance(cell, MergedCell):
                        row_vals.append(
                            {
                                "value": "(merged)",
                                "col": str(cell.column),
                                "row": cell.row,
                            }
                        )
                    else:
                        row_vals.append(
                            {
                                "value": cell.value,
                                "col": cell.column_letter,
                                "row": cell.row,
                            }
                        )
                rows_data.append(row_vals)
            sheets_data[sname] = {
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "merged_cells": merged,
                "rows": rows_data,
            }
        result["sheets_data"] = sheets_data
        wb.close()
    except Exception as e:
        result["error"] = str(e)
    return result


def analyze_pdf(filepath: Path) -> dict:
    result = {
        "type": "pdf",
        "file": filepath.name,
        "size_kb": round(filepath.stat().st_size / 1024, 1),
    }
    try:
        doc = fitz.open(filepath)
        result["page_count"] = len(doc)
        pages = []
        total_text_len = 0
        total_images = 0
        for i, page in enumerate(doc):
            text = page.get_text()
            total_text_len += len(text.strip())
            images = page.get_images(full=True)
            total_images += len(images)
            table_finder = page.find_tables()
            tables_list = table_finder.tables if hasattr(table_finder, "tables") else []
            table_data = []
            for t in tables_list:
                try:
                    extracted = t.extract()
                    table_data.append(extracted[:15])
                except:
                    pass
            pages.append(
                {
                    "page_num": i + 1,
                    "text_length": len(text.strip()),
                    "text_preview": text[:2000] if text.strip() else "(no text)",
                    "table_count": len(tables_list),
                    "tables": table_data,
                    "image_count": len(images),
                }
            )
        result["is_scanned"] = total_text_len < 100
        result["total_text_length"] = total_text_len
        result["total_images"] = total_images
        result["pages"] = pages
        doc.close()
    except Exception as e:
        result["error"] = str(e)
    return result


def main():
    customers = {}
    for item in sorted(INPUT_DIR.iterdir()):
        if item.is_dir():
            files = sorted([f for f in item.rglob("*") if f.is_file()])
            customers[item.name] = files

    all_results = {}
    for customer, files in customers.items():
        print(f"\nAnalyzing: {customer} ({len(files)} files)")
        customer_results = []
        for f in files:
            print(f"  -> {f.relative_to(INPUT_DIR)}")
            if f.suffix.lower() in (".xlsx", ".xls"):
                r = analyze_excel(f)
            elif f.suffix.lower() == ".pdf":
                r = analyze_pdf(f)
            else:
                r = {"type": "unknown", "file": f.name, "suffix": f.suffix}
            r["relative_path"] = str(f.relative_to(INPUT_DIR))
            r["is_target_template"] = "CadCam_Stuecklistenvorlage" in f.name
            customer_results.append(r)
        all_results[customer] = customer_results

    output_dir = BASE / "data" / "analysis"
    output_dir.mkdir(parents=True, exist_ok=True)
    with open(output_dir / "raw_analysis.json", "w", encoding="utf-8") as fp:
        json.dump(all_results, fp, ensure_ascii=False, indent=2, default=str)
    print(f"\nSaved to {output_dir / 'raw_analysis.json'}")


if __name__ == "__main__":
    main()
