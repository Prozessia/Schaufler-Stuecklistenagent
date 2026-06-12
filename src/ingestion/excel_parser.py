"""Excel Parser — parse .xlsx/.xls files into ParsedBOM."""

from __future__ import annotations

import csv
import io
import logging
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

from src.core.models import (
    ExtractionMethod,
    FileFormat,
    ParsedBOM,
    SourceMetadata,
)
from src.core.positions import normalize_position
from src.ingestion.file_router import infer_customer

logger = logging.getLogger(__name__)


def parse_excel(
    filepath: Path | str,
    sheet_name: str | None = None,
    header_row: int | None = None,
    data_start_row: int | None = None,
) -> ParsedBOM:
    """Parse an Excel file into a ParsedBOM.

    Args:
        filepath: Path to the Excel file.
        sheet_name: Sheet to parse. Defaults to first sheet, or "Stückliste" if it exists.
        header_row: 1-based row number for headers. Auto-detected if None.
        data_start_row: 1-based row number where data starts. Auto-detected if None.
    """
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Select sheet
    ws = _select_sheet(wb, sheet_name)

    # Read all rows (skip totally empty)
    all_rows = _read_all_rows(ws)

    if not all_rows:
        wb.close()
        return _empty_result(filepath)

    # Detect header row
    if header_row is None:
        header_row = _detect_header_row(all_rows)

    # Detect data start
    if data_start_row is None:
        data_start_row = _detect_data_start(all_rows, header_row)

    # Extract headers
    headers = _clean_headers(all_rows[header_row - 1])

    # Extract data rows
    rows: list[dict[str, str | None]] = []
    for row_idx in range(data_start_row - 1, len(all_rows)):
        raw = all_rows[row_idx]
        if _is_empty_row(raw):
            continue
        row_dict: dict[str, str | None] = {}
        for col_idx, hdr in enumerate(headers):
            val = raw[col_idx] if col_idx < len(raw) else None
            row_dict[hdr] = _cell_to_str(val)
        rows.append(row_dict)

    wb.close()

    # Excel position anchor (ticket d): find position column and build
    # expected_position_count + raw_pdf_positions from data rows.
    expected_position_count, raw_pdf_positions = _extract_position_anchor(headers, rows)

    return ParsedBOM(
        source=SourceMetadata(
            filename=filepath.name,
            filepath=str(filepath),
            customer=infer_customer(filepath),
            format=FileFormat.EXCEL,
            extraction_method=ExtractionMethod.OPENPYXL,
            extraction_confidence=0.95,
        ),
        headers=headers,
        rows=rows,
        raw_header_rows=[
            _to_str_list(all_rows[i])
            for i in range(header_row - 1, min(header_row, len(all_rows)))
        ],
        expected_position_count=expected_position_count,
        raw_pdf_positions=raw_pdf_positions,
        metadata={
            "sheet_name": ws.title,
            "header_row": header_row,
            "data_start_row": data_start_row,
            "total_sheets": len(wb.sheetnames) if hasattr(wb, "sheetnames") else 1,
        },
    )


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


_BOM_SHEET_KEYWORDS = {"stückliste", "stueckliste", "bom", "parts"}

_BOM_HEADER_KEYWORDS = {
    "pos", "position", "benennung", "bezeichnung", "description",
    "stk", "stück", "menge", "qty", "werkstoff", "material", "zeichn",
}


def _select_sheet(wb: openpyxl.Workbook, sheet_name: str | None):
    """Select the best sheet from the workbook.

    If sheet_name is given, return it directly.  Otherwise score every visible
    sheet and return the one with the highest score (ties → first sheet wins).

    Scoring per sheet:
      +3  if the sheet name contains a BOM keyword (stückliste/bom/parts/…)
      +n  (capped at 8) for each distinct BOM header keyword found anywhere in
          the first 20 rows as a case-insensitive cell-value substring
      +1  if the sheet has > 5 data rows
    Sheets with no rows are ignored.
    """
    if sheet_name:
        return wb[sheet_name]

    best_ws = wb.active
    best_score = -1

    for name in wb.sheetnames:
        ws = wb[name]
        # Skip completely empty sheets
        if ws.max_row is None or ws.max_row == 0:
            continue

        score = 0

        # Name bonus
        name_lower = name.lower()
        if any(kw in name_lower for kw in _BOM_SHEET_KEYWORDS):
            score += 3

        # Header-keyword hits in first 20 rows (cap at 8)
        keyword_hits: set[str] = set()
        row_count = 0
        for row in ws.iter_rows(
            min_row=1, max_row=min(20, ws.max_row or 20), values_only=True
        ):
            has_any = any(v is not None for v in row)
            if has_any:
                row_count += 1
            for cell_val in row:
                if cell_val is None:
                    continue
                cell_str = str(cell_val).lower()
                for kw in _BOM_HEADER_KEYWORDS:
                    if kw in cell_str:
                        keyword_hits.add(kw)
        score += min(len(keyword_hits), 8)

        # Row count bonus
        if row_count > 5:
            score += 1

        if score > best_score:
            best_score = score
            best_ws = ws

    return best_ws


def _read_all_rows(ws) -> list[list]:
    """Read all rows from the worksheet, resolving merged cells.

    Vertical (and horizontal) merged cells: the value lives only in the
    top-left cell of the merge range.  We build a lookup table from all
    merge ranges so that every cell inside a range returns the top-left
    value instead of None.
    """
    # Build merged-cell value lookup: (row, col) → top-left value
    merged_lookup: dict[tuple[int, int], object] = {}
    for merge_range in ws.merged_cells.ranges:
        # The top-left cell carries the actual value
        tl_cell = ws.cell(row=merge_range.min_row, column=merge_range.min_col)
        tl_value = tl_cell.value
        for row_idx in range(merge_range.min_row, merge_range.max_row + 1):
            for col_idx in range(merge_range.min_col, merge_range.max_col + 1):
                merged_lookup[(row_idx, col_idx)] = tl_value

    rows = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), start=1
    ):
        vals = []
        for col_idx, cell in enumerate(row, start=1):
            key = (row_idx, col_idx)
            if isinstance(cell, MergedCell):
                vals.append(merged_lookup.get(key))
            else:
                vals.append(cell.value)
        rows.append(vals)
    return rows


def _detect_header_row(rows: list[list]) -> int:
    """Find the header row — prefer rows where BOM keywords appear.

    Candidate score = text_count + 3 * keyword_hits so that a genuine header
    row beats metadata blocks with many text cells but no column names.
    Falls back to the row with the most text cells when no keyword hits occur.
    """
    best_row = 1
    best_score = 0

    for i, row in enumerate(rows[:20]):  # Check first 20 rows
        text_count = sum(1 for v in row if isinstance(v, str) and len(v.strip()) > 1)
        keyword_hits = 0
        for v in row:
            if v is None:
                continue
            cell_str = str(v).lower()
            for kw in _BOM_HEADER_KEYWORDS:
                if kw in cell_str:
                    keyword_hits += 1
                    break  # one hit per cell is enough
        score = text_count + 3 * keyword_hits
        if score > best_score:
            best_score = score
            best_row = i + 1  # 1-indexed

    return best_row


def _detect_data_start(rows: list[list], header_row: int) -> int:
    """Find first data row after the header — typically header_row + 1 or + 2."""
    for i in range(header_row, min(header_row + 5, len(rows))):
        row = rows[i]
        # Data rows typically have numbers or non-header text
        has_number = any(isinstance(v, (int, float)) for v in row if v is not None)
        non_empty = sum(1 for v in row if v is not None)
        if has_number and non_empty >= 2:
            return i + 1  # 1-indexed
    return header_row + 1


def _clean_headers(raw_row: list) -> list[str]:
    """Convert a raw row into clean header strings."""
    headers = []
    seen: dict[str, int] = {}
    for val in raw_row:
        h = _cell_to_str(val) or ""
        h = h.strip().replace("\n", " ").replace("\r", " ")
        # Remove excessive whitespace
        h = " ".join(h.split())
        if not h:
            h = f"_col_{len(headers) + 1}"

        # De-duplicate
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0

        headers.append(h)
    return headers


def _cell_to_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, str):
        return val.strip() if val.strip() else None
    return str(val)


def _to_str_list(row: list) -> list[str | None]:
    return [_cell_to_str(v) for v in row]


def _is_empty_row(row: list) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row)


_POSITION_COLUMN_KEYWORDS = {"pos", "position", "nr"}


def _find_position_column(headers: list[str]) -> str | None:
    """Return the first header that looks like a position column, or None."""
    for hdr in headers:
        hdr_lower = hdr.lower()
        if any(kw in hdr_lower for kw in _POSITION_COLUMN_KEYWORDS):
            return hdr
    return None


def _extract_position_anchor(
    headers: list[str],
    rows: list[dict[str, str | None]],
) -> tuple[int, list[str]]:
    """Build expected_position_count and raw_pdf_positions from data rows.

    Mirrors the Vision-path logic: order-preserving, deduplicated list of
    normalized position values.  Returns (0, []) when no position column
    is recognisable.
    """
    pos_col = _find_position_column(headers)
    if pos_col is None:
        return 0, []

    seen: set[str] = set()
    positions: list[str] = []
    for row in rows:
        raw_val = row.get(pos_col)
        if raw_val is None:
            continue
        norm = normalize_position(raw_val)
        if not norm:
            continue
        if norm not in seen:
            seen.add(norm)
            positions.append(norm)

    return len(positions), positions


def _empty_result(filepath: Path) -> ParsedBOM:
    return ParsedBOM(
        source=SourceMetadata(
            filename=filepath.name,
            filepath=str(filepath),
            customer=infer_customer(filepath),
            format=FileFormat.EXCEL,
            extraction_method=ExtractionMethod.OPENPYXL,
            extraction_confidence=0.0,
        ),
    )


# ---------------------------------------------------------------------------
# CSV parser
# ---------------------------------------------------------------------------


def parse_csv(filepath: Path | str) -> ParsedBOM:
    """Parse a CSV file into a ParsedBOM.

    Tries UTF-8-with-BOM first, then cp1252 as fallback (German locale).
    Sniffs the dialect; if sniffing fails, tries semicolon then comma.
    Reuses the same header-detection and cleaning helpers as parse_excel.
    """
    filepath = Path(filepath)
    raw_bytes = filepath.read_bytes()

    # Encoding detection: utf-8-sig (handles BOM), then cp1252 fallback.
    text: str | None = None
    used_encoding = "utf-8-sig"
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            text = raw_bytes.decode(encoding)
            used_encoding = encoding
            break
        except UnicodeDecodeError:
            continue

    if text is None:
        # Last resort: replace errors
        text = raw_bytes.decode("utf-8", errors="replace")
        used_encoding = "utf-8"

    # Dialect sniffing
    sample = text[:4096]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        # Fallback: prefer semicolon for German-locale CSVs
        if ";" in sample:
            delimiter = ";"
        else:
            delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows: list[list[str]] = [row for row in reader]

    if not all_rows:
        return _empty_csv_result(filepath, delimiter, used_encoding)

    # Detect header row using the same heuristic as parse_excel
    all_rows_as_lists: list[list] = [list(r) for r in all_rows]
    header_row = _detect_header_row(all_rows_as_lists)
    data_start_row = _detect_data_start(all_rows_as_lists, header_row)

    headers = _clean_headers(all_rows_as_lists[header_row - 1])

    rows: list[dict[str, str | None]] = []
    for row_idx in range(data_start_row - 1, len(all_rows_as_lists)):
        raw = all_rows_as_lists[row_idx]
        if _is_empty_row(raw):
            continue
        row_dict: dict[str, str | None] = {}
        for col_idx, hdr in enumerate(headers):
            val = raw[col_idx] if col_idx < len(raw) else None
            row_dict[hdr] = _cell_to_str(val)
        rows.append(row_dict)

    return ParsedBOM(
        source=SourceMetadata(
            filename=filepath.name,
            filepath=str(filepath),
            customer=infer_customer(filepath),
            format=FileFormat.CSV,
            extraction_method=ExtractionMethod.CSV,
            extraction_confidence=0.95,
        ),
        headers=headers,
        rows=rows,
        raw_header_rows=[
            _to_str_list(all_rows_as_lists[i])
            for i in range(header_row - 1, min(header_row, len(all_rows_as_lists)))
        ],
        metadata={
            "delimiter": delimiter,
            "encoding": used_encoding,
            "header_row": header_row,
            "data_start_row": data_start_row,
        },
    )


def _empty_csv_result(filepath: Path, delimiter: str, encoding: str) -> ParsedBOM:
    return ParsedBOM(
        source=SourceMetadata(
            filename=filepath.name,
            filepath=str(filepath),
            customer=infer_customer(filepath),
            format=FileFormat.CSV,
            extraction_method=ExtractionMethod.CSV,
            extraction_confidence=0.0,
        ),
        metadata={"delimiter": delimiter, "encoding": encoding},
    )
