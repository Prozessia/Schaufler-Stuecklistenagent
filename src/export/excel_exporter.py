"""Excel Exporter — template-based export into the Schaufler Excel template.

Loads the CadCam_Stuecklistenvorlage template, writes scored/transformed
data into the correct cells, and optionally adds traffic-light colour coding
and an audit sheet.

Template layout (V191):
  Row 1:   Title (merged A1:AD1)
  Row 2-3: Meta info (customer, description, drawing number)
  Row 4:   Spacer (height 4.5, usually empty)
  Row 5:   Header row (bold, Arial 10)
  Row 6:   Defaults (AB/AC/AD = "Nein")
  Row 7+:  Data rows (Arial 10, v-align=top, row-height 26.25)
"""

from __future__ import annotations

import copy
import logging
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from src.core.exceptions import ZeroDataLossError
from src.core.positions import POSITION_FIELDS, normalize_position
from src.mapping.schema_registry import TargetSchema, load_schema
from src.scoring.audit_trail import BomAuditTrail, CellAudit
from src.scoring.threshold_manager import TrafficLight

logger = logging.getLogger(__name__)

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
_CONFIG_DIR = _PROJECT_ROOT / "config"

# Traffic-light fill colours (light tints to keep readability)
_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_NEUTRAL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
_FILL_MANUAL = PatternFill(start_color="BFDBFE", end_color="BFDBFE", fill_type="solid")

_FILL_MAP = {
    TrafficLight.GREEN: _FILL_GREEN,
    TrafficLight.YELLOW: _FILL_YELLOW,
    TrafficLight.RED: _FILL_RED,
    TrafficLight.NEUTRAL: _FILL_NEUTRAL,
    TrafficLight.MANUAL_CONFIRMED: _FILL_MANUAL,
}

# Default data-row height matching the template
_DATA_ROW_HEIGHT = 26.25
_DATA_START_ROW = 7  # First data row in template
_SHEET_NAME = "Stückliste"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def export_to_excel(
    audit: BomAuditTrail,
    output_path: Path,
    *,
    template_path: Path | None = None,
    schema: TargetSchema | None = None,
    colour_cells: bool = True,
    add_audit_sheet: bool = True,
    meta: dict[str, str] | None = None,
) -> Path:
    """Export scored BOM data into the Schaufler Excel template.

    Args:
        audit:          Scored BOM audit trail (from ensemble_scorer).
        output_path:    Where to write the output .xlsx.
        template_path:  Path to the template file (default: config/target_template.xlsx).
        schema:         Target schema (loaded from config if None).
        colour_cells:   Whether to apply traffic-light background fills.
        add_audit_sheet: Whether to add a second sheet with full audit details.
        meta:           Optional dict with keys 'order_number', 'customer',
                        'description', 'drawing_number' for meta rows.

    Returns:
        The output_path.
    """
    if template_path is None:
        template_path = _CONFIG_DIR / "target_template.xlsx"
    if schema is None:
        schema = load_schema()

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Load template preserving everything
    wb = openpyxl.load_workbook(template_path)
    ws = wb[_SHEET_NAME]

    # Build column index: field_name -> column letter -> column number
    field_col = {f.name: column_index_from_string(f.column) for f in schema.fields}

    # R3: rows the reviewer deliberately excluded never reach the output. This is an
    # explicit, audited action — the guard below subtracts their identities from the
    # expected set so it is NOT mistaken for silent data loss.
    excluded_rows = set(audit.excluded_rows or [])
    active_cells = [c for c in audit.cells if c.row_index not in excluded_rows]

    # Gather distinct row indices from audit, sorted (excluded rows dropped)
    row_indices = sorted({c.row_index for c in active_cells})

    # A2/ZDL-4: zero-data-loss guard — enforced BEFORE any wb.save().
    # Preferred: SET comparison — every expected position id must appear in the
    # output (catches the case where the count matches but identities differ).
    # Fallback (no position anchor, guard_basis="row_count_fallback"): row-count
    # lower bound. Skipped only when there is genuinely nothing to guard.
    actual_rows = len(row_indices)
    expected_keys = {k for k in (audit.expected_row_keys or []) if k}
    expected_ids = {p for p in (audit.expected_position_ids or []) if p}
    if expected_keys:
        # RB-1 ROW-BAND set guard (deterministic text-layer path): every expected
        # spatial band id must appear in the output, by row identity — catches a
        # dropped row even when it shares a position number (T-007) or has none.
        output_keys = {c.source_row_id for c in active_cells if c.source_row_id}
        # Bands carried ONLY by reviewer-excluded rows are deliberately gone — drop
        # them from the expected set so the guard does not flag an explicit removal.
        excluded_keys = {
            c.source_row_id
            for c in audit.cells
            if c.row_index in excluded_rows and c.source_row_id
        }
        expected_keys -= excluded_keys - output_keys
        missing_keys = sorted(expected_keys - output_keys)
        if missing_keys:
            preview = ", ".join(missing_keys[:10])
            suffix = " …" if len(missing_keys) > 10 else ""
            raise ZeroDataLossError(
                f"DATA LOSS DETECTED: {len(missing_keys)} von {len(expected_keys)} "
                f"erwarteten Zeilen-Bändern fehlen im Output ({preview}{suffix}). "
                f"Datei wird NICHT gespeichert."
            )
    elif expected_ids:
        output_positions = {
            normalize_position(c.transformed_value)
            for c in active_cells
            if c.target_field in POSITION_FIELDS and c.transformed_value
        }
        # Positions carried ONLY by excluded rows are deliberately gone.
        excluded_positions = {
            normalize_position(c.transformed_value)
            for c in audit.cells
            if c.row_index in excluded_rows
            and c.target_field in POSITION_FIELDS
            and c.transformed_value
        }
        expected_ids -= excluded_positions - output_positions
        missing_ids = sorted(expected_ids - output_positions)
        if missing_ids:
            preview = ", ".join(missing_ids[:10])
            suffix = " …" if len(missing_ids) > 10 else ""
            raise ZeroDataLossError(
                f"DATA LOSS DETECTED: {len(missing_ids)} von {len(expected_ids)} "
                f"erwarteten Positionen fehlen im Output ({preview}{suffix}). "
                f"Datei wird NICHT gespeichert."
            )
    elif audit.expected_position_count > 0:
        # Each excluded row was one of the expected positions; subtract them.
        expected_positions = audit.expected_position_count - len(excluded_rows)
        if actual_rows < expected_positions:
            raise ZeroDataLossError(
                f"DATA LOSS DETECTED: {actual_rows} Zeilen im Output, "
                f"aber {expected_positions} Zeilen erwartet "
                f"(guard_basis={audit.guard_basis}). "
                f"Fehlend: {expected_positions - actual_rows}. "
                f"Datei wird NICHT gespeichert."
            )
    else:
        logger.warning(
            "Zero-data-loss guard skipped for %s: no position anchor "
            "(guard_basis=%s)",
            audit.source_file,
            audit.guard_basis,
        )

    if not row_indices:
        logger.warning("No data to export for %s", audit.source_file)
        wb.save(output_path)
        return output_path

    # Clear existing data rows (7 to max_row)
    _clear_data_rows(ws)

    # Store reference formatting from row 7 (first data row) per column
    ref_styles = _capture_row_styles(ws, _DATA_START_ROW)

    # Index audit cells by (row_index, target_field)
    cell_index: dict[tuple[int, str], CellAudit] = {}
    for c in audit.cells:
        cell_index[(c.row_index, c.target_field)] = c

    # Write rows
    for out_row_offset, src_row_idx in enumerate(row_indices):
        excel_row = _DATA_START_ROW + out_row_offset

        # Set row height
        ws.row_dimensions[excel_row].height = _DATA_ROW_HEIGHT

        for field in schema.fields:
            col_num = field_col[field.name]
            cell_audit = cell_index.get((src_row_idx, field.name))

            # Determine value
            value = None
            classification = TrafficLight.RED
            if cell_audit:
                classification = cell_audit.classification
                if cell_audit.transformed_value:
                    value = _coerce_value(cell_audit.transformed_value, field.type)

            ws_cell = ws.cell(row=excel_row, column=col_num, value=value)

            # Apply reference formatting from template
            _apply_style(ws_cell, ref_styles.get(col_num))

            # Apply traffic-light colour
            if colour_cells and cell_audit is not None:
                ws_cell.fill = _FILL_MAP.get(classification, _FILL_RED)

        # Lossless footer/header/note marker: a non-destructive cell comment on
        # the first column. The row stays fully present (zero-data-loss); the
        # comment only flags "this may not be a real table row" for the reviewer.
        non_data_reasons = (audit.non_data_row_flags or {}).get(src_row_idx)
        if non_data_reasons:
            note = ws.cell(row=excel_row, column=1)
            note.comment = Comment(
                "Mögliche Kopf-/Fußzeile oder Notiz – bitte prüfen.\n"
                "Gründe: " + ", ".join(non_data_reasons),
                "BOM-Mapper",
            )

    # Fill meta rows if provided
    if meta:
        _write_meta(ws, meta)

    # Add audit detail sheet
    if add_audit_sheet:
        _add_audit_sheet(wb, audit)

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Exported %d rows to %s", len(row_indices), output_path)
    return output_path


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------


def _clear_data_rows(ws) -> None:
    """Clear cell values in data rows, preserving formatting."""
    for row in ws.iter_rows(
        min_row=_DATA_START_ROW, max_row=ws.max_row, min_col=1, max_col=30
    ):
        for cell in row:
            cell.value = None


def _capture_row_styles(ws, row: int) -> dict[int, dict]:
    """Capture formatting from a reference row for later application."""
    styles: dict[int, dict] = {}
    for col_idx in range(1, 31):
        cell = ws.cell(row=row, column=col_idx)
        styles[col_idx] = {
            "font": copy.copy(cell.font),
            "alignment": copy.copy(cell.alignment),
            "border": copy.copy(cell.border),
            "number_format": cell.number_format,
        }
    return styles


def _apply_style(cell, style_dict: dict | None) -> None:
    """Apply captured style to a cell (font, alignment, border, number_format)."""
    if not style_dict:
        return
    cell.font = style_dict["font"]
    cell.alignment = style_dict["alignment"]
    cell.border = style_dict["border"]
    cell.number_format = style_dict["number_format"]


def _coerce_value(value: str, field_type: str):
    """Coerce string value to appropriate Python type for Excel."""
    if field_type == "integer":
        try:
            return int(value)
        except (ValueError, TypeError):
            return value
    elif field_type == "decimal":
        try:
            return float(value.replace(",", "."))
        except (ValueError, TypeError):
            return value
    return value


def _write_meta(ws, meta: dict[str, str]) -> None:
    """Write meta information into template header rows."""
    if "order_number" in meta:
        ws.cell(row=1, column=1).value = (
            f"Stückliste für Auftrag {meta['order_number']}"
        )
    if "customer" in meta:
        ws.cell(row=3, column=1).value = meta["customer"]
    if "description" in meta:
        ws.cell(row=3, column=7).value = meta["description"]
    if "drawing_number" in meta:
        ws.cell(row=3, column=14).value = meta["drawing_number"]


def _add_audit_sheet(wb: openpyxl.Workbook, audit: BomAuditTrail) -> None:
    """Add a sheet with per-cell audit details."""
    sheet_name = "Audit Trail"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Header
    headers = [
        "Row",
        "Target Field",
        "Column",
        "Raw Value",
        "Transformed Value",
        "Method",
        "Transform Score",
        "Rule Score",
        "Counter Score",
        "Final Score",
        "Classification",
        "Reasoning",
    ]
    header_font = Font(name="Arial", size=10, bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font

    # Data
    data_font = Font(name="Arial", size=9)
    for row_offset, c in enumerate(audit.cells, 2):
        ws.cell(row=row_offset, column=1, value=c.row_index).font = data_font
        ws.cell(row=row_offset, column=2, value=c.target_field).font = data_font
        ws.cell(row=row_offset, column=3, value=c.target_column).font = data_font
        ws.cell(row=row_offset, column=4, value=(c.raw_value or "")[:200]).font = (
            data_font
        )
        ws.cell(
            row=row_offset, column=5, value=(c.transformed_value or "")[:200]
        ).font = data_font
        ws.cell(row=row_offset, column=6, value=c.transform_method).font = data_font
        ws.cell(
            row=row_offset, column=7, value=round(c.transform_confidence, 3)
        ).font = data_font
        ws.cell(row=row_offset, column=8, value=round(c.rule_score, 3)).font = data_font
        cs = c.counter_check_score
        ws.cell(
            row=row_offset, column=9, value=round(cs, 3) if cs is not None else ""
        ).font = data_font
        ws.cell(row=row_offset, column=10, value=round(c.final_score, 3)).font = (
            data_font
        )

        cls_cell = ws.cell(
            row=row_offset, column=11, value=c.classification.value.upper()
        )
        cls_cell.font = data_font
        cls_cell.fill = _FILL_MAP.get(c.classification, _FILL_RED)

        ws.cell(row=row_offset, column=12, value=c.reasoning[:300]).font = data_font

    # Auto-fit approximate column widths
    col_widths = [6, 22, 8, 35, 35, 25, 12, 12, 12, 12, 14, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary row
    summary_row = len(audit.cells) + 3
    ws.cell(row=summary_row, column=1, value="Summary").font = header_font
    ws.cell(row=summary_row, column=2, value=f"GREEN: {audit.green_count}").font = (
        data_font
    )
    ws.cell(row=summary_row, column=3, value=f"YELLOW: {audit.yellow_count}").font = (
        data_font
    )
    ws.cell(row=summary_row, column=4, value=f"RED: {audit.red_count}").font = data_font
    ws.cell(
        row=summary_row,
        column=5,
        value=f"MANUAL: {audit.manual_confirmed_count}",
    ).font = data_font
    ws.cell(row=summary_row, column=6, value=f"Total: {audit.total_scored}").font = (
        data_font
    )
