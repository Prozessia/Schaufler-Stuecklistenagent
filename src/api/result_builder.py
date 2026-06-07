"""Helpers for turning persisted job audits into API result payloads."""

from __future__ import annotations

from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from src.api.job_store import Job
from src.api.models.schemas import (
    CellResult,
    JobResult,
    RowResult,
    TemplateColumnResult,
    TemplateDefaultCellResult,
    TemplateLayoutResult,
    TemplateMetaSectionResult,
)
from src.mapping.schema_registry import TargetField, TargetSchema, load_schema
from src.scoring.threshold_manager import TrafficLight

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
_DEFAULT_TEMPLATE_PATH = _PROJECT_ROOT / "config" / "target_template.xlsx"
_META_SECTION_KEYS = {
    "A": "customer",
    "G": "description",
    "N": "drawing_number",
}

_SEVERITY_RANK = {
    TrafficLight.NEUTRAL: 0,
    TrafficLight.GREEN: 1,
    TrafficLight.MANUAL_CONFIRMED: 1,
    TrafficLight.YELLOW: 2,
    TrafficLight.RED: 3,
}


def excel_column_sort_key(column: str) -> tuple[int, str]:
    """Sort Excel-style columns naturally: A, B, Z, AA, AB ..."""
    normalized = (column or "").strip()
    if not normalized:
        return (10_000, "")

    try:
        return (column_index_from_string(normalized.upper()), normalized)
    except ValueError:
        return (10_000, normalized)


def _fallback_header_lines(field: TargetField) -> list[str]:
    lines = [field.name.strip()] if field.name.strip() else []
    name_de = field.name_de.strip()
    if name_de and name_de != field.name.strip():
        lines.append(name_de)
    return lines or [field.column]


def _split_header_lines(
    header_value: object, field: TargetField
) -> tuple[str, list[str]]:
    label = str(header_value).strip() if header_value not in (None, "") else ""
    if label:
        lines = [line.strip() for line in label.splitlines() if line.strip()]
        if lines:
            return label, lines
    lines = _fallback_header_lines(field)
    return "\n".join(lines), lines


def _build_template_metadata(
    schema: TargetSchema,
    template_path: Path | None = None,
) -> tuple[TemplateLayoutResult, list[TemplateColumnResult]]:
    ordered_fields = sorted(
        schema.fields, key=lambda field: excel_column_sort_key(field.column)
    )
    template_info = schema.template_info
    sheet_name = template_info.sheet or "Stückliste"
    header_row = template_info.header_row or 5
    data_start_row = template_info.data_start_row or 7

    layout = TemplateLayoutResult(
        sheet_name=sheet_name,
        header_row=header_row,
        data_start_row=data_start_row,
    )
    columns: list[TemplateColumnResult] = []

    workbook_path = template_path or _DEFAULT_TEMPLATE_PATH
    if workbook_path.exists():
        workbook = load_workbook(workbook_path, data_only=True)
        worksheet = (
            workbook[sheet_name]
            if sheet_name in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        freeze_panes = worksheet.freeze_panes
        layout = TemplateLayoutResult(
            title=str(worksheet["A1"].value or ""),
            sheet_name=worksheet.title,
            header_row=header_row,
            data_start_row=data_start_row,
            freeze_panes=getattr(freeze_panes, "coordinate", None)
            or (str(freeze_panes) if freeze_panes else None),
            header_height=worksheet.row_dimensions[header_row].height,
            data_row_height=worksheet.row_dimensions[data_start_row].height,
            default_row_height=worksheet.row_dimensions[6].height,
            meta_sections=_build_meta_sections(worksheet),
            default_cells=_build_default_cells(worksheet, schema),
        )

        for field in ordered_fields:
            header_label, header_lines = _split_header_lines(
                worksheet[f"{field.column}{header_row}"].value,
                field,
            )
            sample_cell = worksheet[f"{field.column}{data_start_row}"]
            columns.append(
                TemplateColumnResult(
                    field=field.name,
                    column=field.column,
                    header_label=header_label,
                    header_lines=header_lines,
                    width=worksheet.column_dimensions[field.column].width,
                    type=field.type,
                    required=field.required,
                    horizontal_alignment=sample_cell.alignment.horizontal,
                    vertical_alignment=sample_cell.alignment.vertical,
                )
            )
        return layout, columns

    for field in ordered_fields:
        header_label, header_lines = _split_header_lines(None, field)
        columns.append(
            TemplateColumnResult(
                field=field.name,
                column=field.column,
                header_label=header_label,
                header_lines=header_lines,
                type=field.type,
                required=field.required,
            )
        )

    return layout, columns


def _build_meta_sections(worksheet) -> list[TemplateMetaSectionResult]:
    merged_ranges = list(worksheet.merged_cells.ranges)
    by_start_column: dict[int, dict[int, object]] = {}

    for merged_range in merged_ranges:
        if merged_range.min_row not in {2, 3}:
            continue
        by_start_column.setdefault(merged_range.min_col, {})[
            merged_range.min_row
        ] = merged_range

    sections: list[TemplateMetaSectionResult] = []
    for start_column in sorted(by_start_column):
        label_range = by_start_column[start_column].get(2)
        value_range = by_start_column[start_column].get(3)
        if label_range is None and value_range is None:
            continue

        anchor_col = start_column
        label_cell = worksheet.cell(2, anchor_col)
        value_cell = worksheet.cell(3, anchor_col)
        label = str(label_cell.value or "").strip()
        value = str(value_cell.value or "").strip()
        range_for_bounds = value_range or label_range
        section_column = get_column_letter(
            label_range.min_col if label_range else anchor_col
        )

        sections.append(
            TemplateMetaSectionResult(
                key=_META_SECTION_KEYS.get(section_column, f"section_{anchor_col}"),
                label=label,
                value=value,
                start_column=get_column_letter(range_for_bounds.min_col),
                end_column=get_column_letter(range_for_bounds.max_col),
                label_row=2,
                value_row=3,
                label_horizontal_alignment=label_cell.alignment.horizontal,
                label_vertical_alignment=label_cell.alignment.vertical,
                value_horizontal_alignment=value_cell.alignment.horizontal,
                value_vertical_alignment=value_cell.alignment.vertical,
            )
        )

    return sections


def _build_default_cells(
    worksheet,
    schema: TargetSchema,
) -> list[TemplateDefaultCellResult]:
    default_cells: list[TemplateDefaultCellResult] = []

    for field in sorted(
        schema.fields, key=lambda item: excel_column_sort_key(item.column)
    ):
        worksheet_cell = worksheet[f"{field.column}6"]
        if worksheet_cell.value in (None, ""):
            continue

        default_cells.append(
            TemplateDefaultCellResult(
                field=field.name,
                column=field.column,
                value=str(worksheet_cell.value),
                horizontal_alignment=worksheet_cell.alignment.horizontal,
                vertical_alignment=worksheet_cell.alignment.vertical,
            )
        )

    return default_cells


def _merge_runtime_template_values(
    layout: TemplateLayoutResult,
    job: Job,
) -> TemplateLayoutResult:
    runtime_sections: list[TemplateMetaSectionResult] = []
    for section in layout.meta_sections:
        next_value = section.value
        if section.key == "customer" and job.customer:
            next_value = job.customer
        runtime_sections.append(section.model_copy(update={"value": next_value}))

    return layout.model_copy(update={"meta_sections": runtime_sections})


@lru_cache(maxsize=1)
def _get_default_template_metadata() -> (
    tuple[TargetSchema, TemplateLayoutResult, list[TemplateColumnResult]]
):
    schema = load_schema()
    layout, columns = _build_template_metadata(schema)
    return schema, layout, columns


def build_job_result(
    job_id: str,
    job: Job,
    *,
    schema: TargetSchema | None = None,
    template_path: Path | None = None,
) -> JobResult:
    """Convert a persisted job audit trail into the public JobResult model."""
    if not job.audit:
        raise ValueError("No audit data available")

    if schema is None and template_path is None:
        schema, template_layout, template_columns = _get_default_template_metadata()
    else:
        schema = schema or load_schema()
        template_layout, template_columns = _build_template_metadata(
            schema, template_path
        )

    template_layout = _merge_runtime_template_values(template_layout, job)

    audit = job.audit
    excluded_rows = set(audit.excluded_rows or [])
    rows_dict: dict[int, list] = {}
    first_column_by_field: dict[str, str] = {}

    for cell in audit.cells:
        # Column metadata is derived from all cells (so excluded rows don't change
        # the visible columns), but excluded rows are not rendered as data rows.
        first_column_by_field.setdefault(cell.target_field, cell.target_column)
        if cell.row_index in excluded_rows:
            continue
        rows_dict.setdefault(cell.row_index, []).append(cell)

    known_fields = {column.field for column in template_columns}
    extra_fields = sorted(
        (field for field in first_column_by_field if field not in known_fields),
        key=lambda field: excel_column_sort_key(first_column_by_field[field]),
    )
    for field in extra_fields:
        template_columns.append(
            TemplateColumnResult(
                field=field,
                column=first_column_by_field[field],
                header_label=field,
                header_lines=[field],
            )
        )

    target_fields = [column.field for column in template_columns]
    field_rank = {field: index for index, field in enumerate(target_fields)}

    row_results: list[RowResult] = []
    for row_idx in sorted(rows_dict):
        cells = sorted(
            rows_dict[row_idx],
            key=lambda cell: (
                field_rank.get(cell.target_field, 10_000),
                excel_column_sort_key(cell.target_column),
            ),
        )
        worst = TrafficLight.NEUTRAL
        cell_results: list[CellResult] = []

        for cell in cells:
            cell_results.append(
                CellResult(
                    row_index=cell.row_index,
                    target_field=cell.target_field,
                    target_column=cell.target_column,
                    raw_value=cell.raw_value,
                    transformed_value=cell.transformed_value,
                    method=cell.transform_method,
                    score=round(cell.final_score, 3),
                    classification=cell.classification.value,
                    final_status=cell.final_status,
                    value_match_result=cell.value_match_result,
                    value_match_detail=cell.value_match_detail,
                    field_category=cell.field_category,
                    green_evidence=list(cell.green_evidence),
                    blocking_errors=list(cell.blocking_errors),
                    hard_vetoes=list(cell.hard_vetoes),
                    reasoning=cell.reasoning,
                    source_location=cell.source_location,
                )
            )

            if _SEVERITY_RANK[cell.classification] > _SEVERITY_RANK[worst]:
                worst = cell.classification

        non_data_reasons = (audit.non_data_row_flags or {}).get(row_idx, [])
        row_results.append(
            RowResult(
                row_index=row_idx,
                cells=cell_results,
                worst_classification=worst.value,
                non_data=bool(non_data_reasons),
                non_data_reasons=list(non_data_reasons),
            )
        )

    return JobResult(
        job_id=job_id,
        filename=job.filename,
        customer=job.customer,
        total_rows=len(row_results),
        total_cells=audit.total_cells,
        green_count=audit.green_count,
        yellow_count=audit.yellow_count,
        red_count=audit.red_count,
        neutral_count=audit.neutral_count,
        manual_confirmed_count=audit.manual_confirmed_count,
        green_pct=round(audit.green_pct, 1),
        yellow_pct=round(audit.yellow_pct, 1),
        red_pct=round(audit.red_pct, 1),
        neutral_pct=round(audit.neutral_pct, 1),
        manual_confirmed_pct=round(audit.manual_confirmed_pct, 1),
        completeness_guaranteed=audit.completeness_guaranteed,
        completeness_reason=audit.completeness_reason,
        expected_position_count=audit.expected_position_count,
        guard_basis=audit.guard_basis,
        excluded_rows=sorted(excluded_rows),
        target_fields=target_fields,
        columns=template_columns,
        template=template_layout,
        rows=row_results,
    )
